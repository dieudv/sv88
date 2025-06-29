import json
import pandas as pd
import requests
import os
import re
from datetime import datetime, timedelta
from openpyxl import load_workbook
from datetime import timezone

API_URL = "https://be.sb21.net/api/v2/getEvent?timeRange=today&sportType=1_1&sportId=1&oddsStyle=ma&pinLeague=false"
HEADERS = {
    "accept": "application/json",
    "content-type": "application/json",
    "lng": "vi"
}

# Tạo thư mục xuất
os.makedirs("matches", exist_ok=True)

def append_to_excel(filepath, df):
    from openpyxl import load_workbook

    if not os.path.exists(filepath):
        # File chưa tồn tại → ghi mới
        df.to_excel(filepath, index=False)
    else:
        if file_has_final(filepath):
            print(f"Bỏ qua vì đã có dòng 'Chung cuộc' trong {filepath}")
            return
        
        # Mở workbook và tính số dòng đang có
        book = load_workbook(filepath)
        sheet = book.active
        start_row = sheet.max_row

        # Ghi thêm từ dòng tiếp theo
        with pd.ExcelWriter(filepath, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
            df.to_excel(writer, index=False, header=False, startrow=start_row)

def file_has_final(filepath):
    try:
        df_old = pd.read_excel(filepath)
        return "Chung cuộc" in df_old["Thời điểm"].values
    except:
        return False

def sanitize_filename(s):
    return re.sub(r'[\\/*?:"<>|]', "_", s).replace(" ", "_")

def parse_utc_time(iso_string):
    return datetime.strptime(iso_string, "%Y-%m-%dT%H:%M:%SZ").replace(tzinfo=timezone.utc)

def get_time_and_score(match_time_str, match):
    match_time = parse_utc_time(match_time_str)
    now = datetime.now(timezone.utc)
    delta = match_time - now

    if delta > timedelta(minutes=15):
        return now.strftime("%H:%M"), "-"
    elif timedelta(minutes=0) <= delta <= timedelta(minutes=15):
        return "Trước trận", "-"
    elif delta > timedelta(minutes=-150):
        # Trận đang diễn ra → tính phút
        minutes_played = int((now - match_time).total_seconds() // 60)
        hours = minutes_played // 60
        minutes = minutes_played % 60
        time_str = f"{hours}H {minutes}'" if hours else f"{minutes}'"
        score = match.get("4", {})
        return time_str, f"{score.get('0', 0)}-{score.get('1', 0)}"
    else:
        score = match.get("4", {})
        return "Chung cuộc", f"{score.get('0', 0)}-{score.get('1', 0)}"

def extract_odds(odds_list):
    if not odds_list:
        return None, None, None
    try:
        text = odds_list[0]
        # Regex tìm giá trị + hậu tố
        matches = re.findall(r'([\d\.]+)\*\d+h|([\d\.]+)\*\d+a|([\d\.]+)\*\d+d', text)
        home = away = draw = None
        for h, a, d in matches:
            if h:
                home = h
            if a:
                away = a
            if d:
                draw = d
        return home, away, draw
    except Exception as e:
        print(f"Lỗi extract_odds: {e}")
        return None, None, None

try:
    response = requests.get(API_URL, headers=HEADERS)
    response.raise_for_status()
    competitions = response.json()[0]
except Exception as e:
    print(f"[ERROR] Không thể lấy dữ liệu từ API: {e}")    

for comp in competitions:
    comp_name = comp.get("1", "UnknownLeague")
    matches = comp.get("2", [])

    for match in matches:
        # Bỏ kèo phụ
        if "16" in match or match.get("17", False):
            continue

        try:
            time_label, score = get_time_and_score(match.get("0", ""), match)

            home = match.get("2", "Home")
            away = match.get("3", "Away")
            odds = match.get("7", {})

            odds_1x2 = extract_odds(odds.get("1", []))
            ou = extract_odds(odds.get("3", []))
            hc = extract_odds(odds.get("5", []))

            row = {
                "Thời điểm": time_label,
                "Tỉ số": score,
                "Kèo chấp": hc[0],
                "Odds Real": hc[1],
                "Odds Pachuca": hc[2],
                "T/X": ou[0],
                "Odds Tài": ou[1],
                "Odds Xỉu": ou[2],
                "1X2 - Real": odds_1x2[0],
                "1X2 - Hòa": odds_1x2[2],
                "1X2 - Pachuca": odds_1x2[1]
            }

            df = pd.DataFrame([row])
            filename = sanitize_filename(f"{comp_name} - {home} vs {away}.xlsx")
            filepath = os.path.join("matches", filename)
            append_to_excel(filepath, df)

        except Exception as e:
            print(f"Lỗi trận {match.get('2')} vs {match.get('3')}: {e}")
